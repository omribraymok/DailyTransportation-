from clCar import Car
from clHuman import Child
from clSchool import School
import global_variables as gv

# Reading and Writing from/to excel file
import openpyxl
# Calculation of random points
import random

import numpy as np


# this func will divide dictionary
def div_groups(dict_of_all: dict, num_of_parts: int):
    list_len: int = len(gv.dict_of_all_children)
    return [dict(list(dict_of_all.items())[k * list_len // num_of_parts:(k + 1) * list_len // num_of_parts])
            for k in range(num_of_parts)]


# Enter a random point to address column in Child table
def enter_random_point(number):
    for t in range(number):
        temp = "D" + str(t + 2)
        (number_row, y) = random.randrange(-10, 10), random.randrange(-10, 10)
        child_sheet[temp] = str(number_row) + ',' + str(y)

    # Increases the  child table
    tmp = [ws.tables for ws in excel_file.worksheets]
    tables_in_sheet = [{v.name: v} for t in tmp for v in t.values()]

    for table in tables_in_sheet:
        if list(table.keys())[0] == 'Child':
            temp_str = 'A1:G' + str(t + 2)
            table['Child'].ref = temp_str


# Function to print all the point and all the time to excel
def print_matrix_to_excel():
    # create workbook
    wb = openpyxl.Workbook()
    # get worksheet
    ws = wb.active
    # print the point on the first row
    for x in range(2, number_of_children + 3):
        # get a pointer for tab in table
        tab = ws.cell(row=x, column=1)
        # write in the tab
        tab.value = str(gv.point_list[x - 2])
    # print the point on the first row column
    r = 2  # for column use
    for k in gv.dict_of_all_children:
        tab = ws.cell(row=1, column=r)
        r = r + 1
        # write in the tab
        tab.value = str(gv.dict_of_all_children[k])
    # print the time travel matrix
    for x in range(2, number_of_children + 3):
        for y in range(2, number_of_children + 2):
            # get a pointer for tab in table
            tab = ws.cell(row=x, column=y)
            # write in the tab
            tab.value = str(gv.time_matrix[x - 2][y - 2])
    wb.save("matrix.xlsx")


file_number = 0


def print_to_excel_time_cost_per_group(dic, car, cost, time):
    # create workbook
    wb = openpyxl.Workbook()
    # get worksheet
    ws = wb.active
    # # change sheet name
    # ws.title = "Group" + str(sheet_number)
    r = 1  # for column use
    for t in dic:
        # get a pointer for tab in table
        tab = ws.cell(row=1, column=r)
        r = r + 1
        # write in the tab
        tab.value = str(dic[t])

    tab = ws.cell(row=2, column=1)
    # write in the tab
    tab.value = "car id: " + str(car.ID)
    tab = ws.cell(row=3, column=1)
    # write in the tab
    tab.value = "cost: " + str(cost)
    tab = ws.cell(row=4, column=1)
    # write in the tab
    tab.value = "time: " + str(time)
    wb.save("Groups" + str(file_number) + ".xlsx")


def calculate_time_cost_per_group(dic_of_children, car, school_address):
    total_cost = car.driver_cost
    total_time = 0
    for key in dic_of_children:
        # to break when loop get to the obj before the lest one
        if key == ((len(dic_of_children) - 1) * (file_number + 1)):
            break
        (cost, time) = car.calculate_cost_time(dic_of_children[key].address, dic_of_children[key + 1].address)
        total_cost += cost
        total_time += time
    # calculate the time and cost for school
    (cost, time) = car.calculate_cost_time(dic_of_children[key].address, school_address)
    total_cost += cost
    total_time += time
    print_to_excel_time_cost_per_group(dic_of_children, car, total_cost, total_time)


# Using openpyxl to writing to excel file
# Give the location of the file
data_file = "Worksheet.xlsx"

# To open Workbook
excel_file = openpyxl.load_workbook(data_file)

# To open sheets
child_sheet = excel_file.worksheets[0]

number_of_children = 21
enter_random_point(number_of_children)

excel_file.save(data_file)


# get data from excel file
for number_row in range(2, number_of_children + 2):
    gv.dict_of_all_children[number_row - 2] = Child(child_sheet, number_row)

# Reading from excel file from Car table
cars_sheet = excel_file.worksheets[2]
dict_of_cars = {}
# get cars data from data sheet
for number_row in range(2, 5):
    dict_of_cars[number_row - 2] = Car(cars_sheet, number_row)

# Reading from excel file from School table
schools_sheet = excel_file.worksheets[4]
dict_of_school = {}
for number_row in range(2, 3):
    dict_of_school[number_row - 2] = School(schools_sheet, number_row)


# Enter the school's address to the matrix
temp_point = dict_of_school[0].address
(x1, y1) = temp_point.split(',')
(x1, y1) = (int(x1), int(y1))
gv.point_list.insert(0, (x1, y1))
# Enter the children's address to the list
for number_row in range(0, number_of_children):
    temp_point = gv.dict_of_all_children[number_row].address
    (x1, y1) = temp_point.split(',')
    (x1, y1) = (int(x1), int(y1))
    gv.point_list.insert(number_row, (x1, y1))
# Matrix 21X21, this matrix will contain all the time travel from point A (row0) to point B (column0)
gv.time_matrix = np.zeros((number_of_children + 1, number_of_children))
# Enter all the time travel
for key in gv.dict_of_all_children:
    for number_row in range(0, number_of_children + 1):
        gv.time_matrix[number_row, key] = Child.calculate_euclidean_dist(gv.dict_of_all_children[key], gv.point_list[number_row])

print_matrix_to_excel()

for number_row in range(0, 3):
    temp_dict = div_groups(gv.dict_of_all_children, 3)[number_row]
    file_number = number_row
    calculate_time_cost_per_group(temp_dict, dict_of_cars[number_row], dict_of_school[0].address)
